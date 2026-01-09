import psycopg2
from psycopg2.extras import RealDictCursor
import csv
import re
import os
import sys
import subprocess
import time
import socket
from dotenv import load_dotenv  # <--- IMPORTANTE
from contextlib import contextmanager
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Carrega as variáveis do arquivo .env
load_dotenv()

# --- CONFIGURAÇÃO DAS CONEXÕES VIA ENV ---
DB_GESTAO = {
    'host': os.getenv('DB_GESTAO_HOST'),
    'database': os.getenv('DB_GESTAO_NAME'),
    'user': os.getenv('DB_GESTAO_USER'),
    'password': os.getenv('DB_GESTAO_PASS')
}

DB_CONTRATO = {
    'host': os.getenv('DB_CONTRATO_HOST'),
    'database': os.getenv('DB_CONTRATO_NAME'),
    'user': os.getenv('DB_CONTRATO_USER'),
    'password': os.getenv('DB_CONTRATO_PASS')
}

DB_PESSOA = {
    'host': os.getenv('DB_PESSOA_HOST'),
    'database': os.getenv('DB_PESSOA_NAME'),
    'user': os.getenv('DB_PESSOA_USER'),
    'password': os.getenv('DB_PESSOA_PASS')
}

SENHA_ACCOUNTS = os.getenv('DB_ACCOUNTS_PASS')

# --- CONFIGURAÇÃO SSH TUNNEL (OBRIGATÓRIO) ---
SSH_CONFIG = {
    'ssh_host': os.getenv('SSH_HOST'),
    'ssh_user': os.getenv('SSH_USER'),
    'ssh_port': int(os.getenv('SSH_PORT', '22')),
    'ssh_password': os.getenv('SSH_PASSWORD'),
    'ssh_pkey': os.getenv('SSH_PKEY_PATH'),  # Caminho para chave privada (opcional)
    'remote_bind_address': (os.getenv('SSH_REMOTE_DB_HOST', 'localhost'), int(os.getenv('SSH_REMOTE_DB_PORT', '5432'))),
    'local_bind_port': int(os.getenv('SSH_LOCAL_PORT', '5435'))
}

# Validação: túnel SSH é obrigatório
if not SSH_CONFIG['ssh_host'] or not SSH_CONFIG['ssh_user']:
    print("ERRO: SSH_HOST e SSH_USER são obrigatórios no arquivo .env")
    sys.exit(1)

# --- FUNÇÕES AUXILIARES ---
def limpar_cpf(cpf):
    """Remove caracteres não numéricos."""
    if not cpf: return None
    return re.sub(r'\D', '', str(cpf))

def formatar_cpf(cpf):
    """Aplica máscara de CPF (apenas para exibição se necessário)."""
    c = limpar_cpf(cpf)
    if len(c) != 11: return c
    return f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}"

def salvar_csv(nome_arquivo, dados, cabecalho):
    """Salva lista de dicionários em CSV na raiz."""
    caminho = os.path.join(os.getcwd(), nome_arquivo)
    if not dados:
        print(f"-> Arquivo {nome_arquivo} não gerado (sem dados).")
        return

    try:
        with open(caminho, mode='w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=cabecalho)
            writer.writeheader()
            writer.writerows(dados)
        print(f"-> Relatório salvo: {caminho} ({len(dados)} registros)")
    except Exception as e:
        print(f"Erro ao salvar {nome_arquivo}: {e}")

def salvar_excel_consolidado(relatorios_dict, nome_arquivo='relatorio_divergencias.xlsx'):
    """
    Salva múltiplos relatórios em um único arquivo Excel com abas separadas.
    
    Args:
        relatorios_dict: dict com formato {'Nome da Aba': (dados, cabecalho)}
        nome_arquivo: nome do arquivo Excel a ser gerado
    """
    caminho = os.path.join(os.getcwd(), nome_arquivo)
    
    try:
        wb = Workbook()
        # Remove a aba padrão criada
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        for nome_aba, (dados, cabecalho) in relatorios_dict.items():
            # Cria nova aba
            ws = wb.create_sheet(title=nome_aba)
            
            # Se não houver dados, adiciona apenas cabeçalho e mensagem
            if not dados:
                ws.append(cabecalho)
                ws.append(['Nenhum registro encontrado'])
                continue
            
            # Adiciona cabeçalho
            ws.append(cabecalho)
            
            # Estiliza cabeçalho
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col_num, _ in enumerate(cabecalho, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Adiciona dados
            for item in dados:
                linha = [item.get(col, '') for col in cabecalho]
                ws.append(linha)
            
            # Ajusta largura das colunas
            for col_num, col_name in enumerate(cabecalho, 1):
                column_letter = get_column_letter(col_num)
                # Calcula largura baseada no conteúdo
                max_length = len(str(col_name))
                for row in ws.iter_rows(min_row=2, max_row=min(100, len(dados)+1), min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 50)  # Limite de 50 caracteres
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Congela primeira linha (cabeçalho)
            ws.freeze_panes = 'A2'
        
        # Salva o arquivo
        wb.save(caminho)
        
        # Conta total de registros
        total_registros = sum(len(dados) for dados, _ in relatorios_dict.values())
        print(f"\n📊 Relatório Excel consolidado salvo: {caminho}")
        print(f"   └─ {len(relatorios_dict)} abas criadas | {total_registros} registros totais")
        
    except Exception as e:
        print(f"⚠️  Erro ao salvar arquivo Excel: {e}")
        print(f"   Os arquivos CSV individuais foram mantidos como backup.")

def verificar_porta_disponivel(port):
    """Verifica se uma porta está disponível para uso."""
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        sock.bind(('127.0.0.1', port))
        sock.close()
        return True
    except OSError:
        return False

def aguardar_porta_aberta(port, timeout=10):
    """Aguarda até que a porta esteja aberta e aceitando conexões."""
    inicio = time.time()
    while time.time() - inicio < timeout:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        try:
            sock.connect(('127.0.0.1', port))
            sock.close()
            return True
        except (socket.error, ConnectionRefusedError):
            time.sleep(0.5)
    return False

@contextmanager
def gerenciar_tunnel_ssh():
    """Context manager para gerenciar ciclo de vida do túnel SSH usando comando nativo."""
    processo_ssh = None
    
    try:
        print(f"[SSH] Conectando ao servidor {SSH_CONFIG['ssh_host']}:{SSH_CONFIG['ssh_port']}...")
        
        # Verifica se a porta local está disponível
        if not verificar_porta_disponivel(SSH_CONFIG['local_bind_port']):
            print(f"[SSH] Aviso: Porta {SSH_CONFIG['local_bind_port']} já está em uso.")
            print(f"[SSH] Assumindo que o túnel já está ativo...")
            yield None
            return
        
        # Monta o comando SSH para criar o túnel
        # Formato: ssh -L local_port:remote_host:remote_port user@ssh_host -p ssh_port -N -o StrictHostKeyChecking=no
        remote_host, remote_port = SSH_CONFIG['remote_bind_address']
        
        ssh_cmd = [
            'ssh',
            '-L', f"{SSH_CONFIG['local_bind_port']}:{remote_host}:{remote_port}",
            '-p', str(SSH_CONFIG['ssh_port']),
            '-l', SSH_CONFIG['ssh_user'],
            SSH_CONFIG['ssh_host'],
            '-N',  # Não executa comando remoto
            '-o', 'StrictHostKeyChecking=no',  # Aceita host automaticamente
            '-o', 'ServerAliveInterval=60',  # Keep-alive
            '-o', 'ServerAliveCountMax=3'
        ]
        
        # Se houver chave privada, adiciona ao comando
        if SSH_CONFIG['ssh_pkey']:
            ssh_cmd.insert(1, '-i')
            ssh_cmd.insert(2, SSH_CONFIG['ssh_pkey'])
        
        # Inicia o processo SSH em background
        print(f"[SSH] Estabelecendo túnel: localhost:{SSH_CONFIG['local_bind_port']} -> {remote_host}:{remote_port}")
        
        # No Windows, usa CREATE_NEW_PROCESS_GROUP para poder encerrar depois
        if os.name == 'nt':  # Windows
            processo_ssh = subprocess.Popen(
                ssh_cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                stdin=subprocess.PIPE,
                creationflags=subprocess.CREATE_NEW_PROCESS_GROUP
            )
        else:  # Linux/Mac
            processo_ssh = subprocess.Popen(
                ssh_cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                stdin=subprocess.PIPE,
                preexec_fn=os.setsid
            )
        
        # Se houver senha, envia via stdin (funciona com sshpass ou similar)
        if SSH_CONFIG['ssh_password'] and not SSH_CONFIG['ssh_pkey']:
            print("[SSH] Nota: Para autenticação por senha, considere usar chave SSH.")
            print("[SSH] Você pode precisar digitar a senha manualmente...")
        
        # Aguarda o túnel ficar disponível
        print(f"[SSH] Aguardando túnel ficar ativo...", end=" ")
        if aguardar_porta_aberta(SSH_CONFIG['local_bind_port'], timeout=15):
            print("✓")
            print(f"[SSH] Túnel SSH estabelecido com sucesso!")
        else:
            raise Exception("Timeout ao aguardar túnel SSH ficar ativo")
        
        yield processo_ssh
        
    except FileNotFoundError:
        print(f"[SSH] ERRO: Comando 'ssh' não encontrado no sistema.")
        print(f"[SSH] Certifique-se de que o OpenSSH está instalado:")
        print(f"[SSH]   - Windows: Settings > Apps > Optional Features > OpenSSH Client")
        print(f"[SSH]   - Linux: sudo apt-get install openssh-client")
        sys.exit(1)
    except Exception as e:
        print(f"[SSH] Erro ao estabelecer túnel: {e}")
        print("[SSH] Verifique as configurações SSH no arquivo .env")
        if processo_ssh:
            try:
                processo_ssh.terminate()
            except:
                pass
        sys.exit(1)
    finally:
        if processo_ssh:
            print("[SSH] Encerrando túnel SSH...", end=" ")
            try:
                processo_ssh.terminate()
                processo_ssh.wait(timeout=5)
                print("✓")
            except:
                try:
                    processo_ssh.kill()
                    print("✓ (forçado)")
                except:
                    print("⚠️  (processo pode continuar em background)")
            print("[SSH] Túnel SSH encerrado.")

def ajustar_hosts_para_tunnel(db_config):
    """Ajusta host e porta dos bancos para usar túnel SSH (obrigatório)."""
    config = db_config.copy()
    # Todos os bancos acessam via localhost na porta do túnel
    config['host'] = '127.0.0.1'
    config['port'] = SSH_CONFIG['local_bind_port']
    return config

def testar_conexoes(db_gestao, db_contrato, db_pessoa):
    """Testa todas as conexões de banco de dados."""
    print("\n" + "="*50)
    print("TESTANDO CONEXÕES COM OS BANCOS DE DADOS")
    print("="*50)
    
    erros = []
    
    # Teste 1: Banco GESTÃO
    try:
        print("[1/3] Testando conexão com banco GESTÃO...", end=" ")
        conn = psycopg2.connect(**db_gestao)
        conn.close()
        print("✓ OK")
    except Exception as e:
        print("✗ FALHOU")
        erros.append(f"GESTÃO: {e}")
    
    # Teste 2: Banco CONTRATO
    try:
        print("[2/3] Testando conexão com banco CONTRATO...", end=" ")
        conn = psycopg2.connect(**db_contrato)
        conn.close()
        print("✓ OK")
    except Exception as e:
        print("✗ FALHOU")
        erros.append(f"CONTRATO: {e}")
    
    # Teste 3: Banco PESSOA
    try:
        print("[3/3] Testando conexão com banco PESSOA...", end=" ")
        conn = psycopg2.connect(**db_pessoa)
        conn.close()
        print("✓ OK")
    except Exception as e:
        print("✗ FALHOU")
        erros.append(f"PESSOA: {e}")
    
    print("="*50)
    
    if erros:
        print("\n❌ ERRO: Falha ao conectar nos seguintes bancos:")
        for erro in erros:
            print(f"   - {erro}")
        print("\nVerifique as configurações no arquivo .env e tente novamente.")
        return False
    else:
        print("\n✅ Conexões estabelecidas com sucesso!")
        return True

def main():
    print("--- INICIANDO DIAGNÓSTICO DE DIVERGÊNCIAS ---")
    
    # Gerencia túnel SSH automaticamente
    with gerenciar_tunnel_ssh():
        # Ajusta configurações dos bancos para usar túnel se necessário
        db_gestao_ajustado = ajustar_hosts_para_tunnel(DB_GESTAO)
        db_contrato_ajustado = ajustar_hosts_para_tunnel(DB_CONTRATO)
        db_pessoa_ajustado = ajustar_hosts_para_tunnel(DB_PESSOA)
        
        # PASSO 0: TESTE DE CONEXÕES
        conexoes_ok = testar_conexoes(db_gestao_ajustado, db_contrato_ajustado, db_pessoa_ajustado)
        
        if not conexoes_ok:
            print("\n⚠️  Encerrando script devido a erros de conexão.")
            return
        
        # Solicita confirmação do usuário
        print("\n" + "="*50)
        resposta = input("Prosseguir com análise? (S/N): ").strip().upper()
        print("="*50)
        
        if resposta not in ['S', 'SIM', 'Y', 'YES']:
            print("\n⚠️  Análise cancelada pelo usuário.")
            return
        
        print("\n🚀 Iniciando análise de divergências...\n")
        
        # LISTAS PARA RELATÓRIOS
        lista_email_duplicado = []
        lista_um_inexistente = []
        lista_ambos_inexistentes = []
        lista_erros_outros = []

        # 1. BUSCAR DIVERGÊNCIAS (GESTAO + ACCOUNTS via DBLINK)
        print("[1/4] Buscando divergências iniciais...")
        divergencias = []
        
        sql_base = f"""
        WITH divergencias AS (
            SELECT
               a.id AS id_account,
               s.sso_id AS sso_id_gestao,
               a.cpf_cnpj AS cpf_visual_accounts,
               s.cpf_cnpj AS cpf_visual_gestao,
               REGEXP_REPLACE(a.cpf_cnpj, '\D','', 'g') AS cpf_accounts_limpo, 
               REGEXP_REPLACE(s.cpf_cnpj, '\D','', 'g') AS cpf_gestao_limpo
            FROM tb_usuario s
            INNER JOIN (
              SELECT cpf_cnpj, id
              FROM dblink(
               'host=issec-live-db.c9aok84ka6e0.sa-east-1.rds.amazonaws.com dbname=accounts_api user=accounts_api password={SENHA_ACCOUNTS}',
                  'SELECT cpf_cnpj, id FROM users'
              ) AS accounts(cpf_cnpj varchar(255), id uuid) 
            ) a ON s.sso_id = a.id 
            WHERE REGEXP_REPLACE(s.cpf_cnpj,'\D','', 'g') <> REGEXP_REPLACE(a.cpf_cnpj,'\D','', 'g')
        )
        SELECT * FROM divergencias;
        """

        try:
            conn = psycopg2.connect(**db_gestao_ajustado)
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute(sql_base)
            divergencias = cur.fetchall()
            conn.close()
        except Exception as e:
            print(f"Erro crítico ao buscar divergências: {e}")
            return

        if not divergencias:
            print("Nenhuma divergência encontrada. Encerrando.")
            return

        # Coletar todos os CPFs únicos para as próximas consultas (Otimização)
        todos_cpfs = set()
        for d in divergencias:
            todos_cpfs.add(d['cpf_accounts_limpo'])
            todos_cpfs.add(d['cpf_gestao_limpo'])
        
        cpfs_tuple = tuple(todos_cpfs)

        # 2. VERIFICAR EXISTÊNCIA NO CONTRATO (SEGURADO)
        print(f"[2/4] Validando {len(todos_cpfs)} CPFs na tabela Segurado...")
        cpfs_existentes_segurado = set()
        
        try:
            conn = psycopg2.connect(**db_contrato_ajustado)
            cur = conn.cursor()
            # Busca CPFs limpos da tabela segurado que coincidem com nossa lista
            sql_segurado = f"""
                SELECT REGEXP_REPLACE(cpf_cnpj, '\D','', 'g') 
                FROM segurado 
                WHERE REGEXP_REPLACE(cpf_cnpj, '\D','', 'g') IN %s
            """
            cur.execute(sql_segurado, (cpfs_tuple,))
            results = cur.fetchall()
            for row in results:
                cpfs_existentes_segurado.add(row[0]) # Adiciona ao Set de existência
            conn.close()
        except Exception as e:
            print(f"Erro ao consultar Segurado: {e}")
            return

        # 3. BUSCAR EMAILS (PESSOA/CONTATO)
        print("[3/4] Buscando e-mails no quarto banco...")
        mapa_emails = {} # { 'cpf_limpo': 'email' }
        
        try:
            conn = psycopg2.connect(**db_pessoa_ajustado)
            cur = conn.cursor()
            
            # Query solicitada adaptada para buscar em lote
            # Precisamos buscar pelo CPF formatado ou limpo? 
            # Assumindo que o banco pessoa armazena formatado, aplicamos replace no WHERE
            sql_emails = f"""
                SELECT REGEXP_REPLACE(p.cpf_cnpj, '\D','', 'g') as cpf, c.valor as email
                FROM pessoa p 
                LEFT JOIN contato c ON p.id = c.pessoa_id
                WHERE c.tipo = 'EMAIL'
                AND REGEXP_REPLACE(p.cpf_cnpj, '\D','', 'g') IN %s
            """
            cur.execute(sql_emails, (cpfs_tuple,))
            results = cur.fetchall()
            
            for cpf, email in results:
                if email:
                    mapa_emails[cpf] = email.strip() # Normaliza email
            conn.close()
        except Exception as e:
            print(f"Erro ao consultar Emails: {e}")
            return

        # 4. PROCESSAMENTO LÓGICO E CONTAGEM
        print("[4/4] Processando regras de negócio...")

        for item in divergencias:
            cpf_acc = item['cpf_accounts_limpo']
            cpf_ges = item['cpf_gestao_limpo']
            
            # Verifica existência na tabela segurado
            existe_acc = cpf_acc in cpfs_existentes_segurado
            existe_ges = cpf_ges in cpfs_existentes_segurado
            
            # Estrutura base do relatório
            linha_relatorio = {
                'id_accounts': item['id_account'],
                'sso_id_gestao': item['sso_id_gestao'],
                'cpf_gestao': item['cpf_visual_gestao'],
                'cpf_accounts': item['cpf_visual_accounts'],
                'existe_segurado_gestao': "SIM" if existe_ges else "NAO",
                'existe_segurado_accounts': "SIM" if existe_acc else "NAO",
                'dono_real_eh_accounts': "SIM" if existe_acc else "NAO", # Conforme solicitado
                'email_comum': None
            }

            # CASO 1: AMBOS INEXISTENTES EM SEGURADO
            if not existe_acc and not existe_ges:
                lista_ambos_inexistentes.append(linha_relatorio)
                continue

            # CASO 2: UM DELES INEXISTENTE
            if (existe_acc and not existe_ges) or (not existe_acc and existe_ges):
                lista_um_inexistente.append(linha_relatorio)
                # Nota: O fluxo pode parar aqui ou continuar para verificar email mesmo assim?
                # Pela lógica comum, se um não existe, é inconsistência de cadastro, mas vamos verificar email apenas se ambos existirem ou se solicitado.
                # Vou assumir que se falta um, já cai nessa categoria e encerra.
                continue 

            # CASO 3: AMBOS EXISTEM -> VERIFICAR EMAIL
            email_acc = mapa_emails.get(cpf_acc)
            email_ges = mapa_emails.get(cpf_ges)

            if email_acc and email_ges and (email_acc == email_ges):
                linha_relatorio['email_comum'] = email_acc
                lista_email_duplicado.append(linha_relatorio)
            else:
                # Se chegou aqui: existem no segurado, mas emails diferentes ou nulos
                lista_erros_outros.append(linha_relatorio)

        # 5. EXIBIÇÃO E SALVAMENTO
        print("\n" + "="*40)
        print("RESUMO FINAL DA OPERAÇÃO")
        print("="*40)
        print(f"1. E-mails Duplicados (Inconsistência Confirmada): {len(lista_email_duplicado)}")
        print(f"2. Um dos CPFs não existe em Segurado:             {len(lista_um_inexistente)}")
        print(f"3. Ambos CPFs não existem em Segurado:             {len(lista_ambos_inexistentes)}")
        print(f"4. Outros (Existem mas e-mail não bate/nulo):      {len(lista_erros_outros)}")
        print("-" * 40)
        print(f"TOTAL ANALISADO: {len(divergencias)}")
        print("="*40)

        # Headers para CSV
        headers = ['id_accounts', 'sso_id_gestao', 'cpf_gestao', 'cpf_accounts', 
                   'existe_segurado_gestao', 'existe_segurado_accounts', 
                   'dono_real_eh_accounts', 'email_comum']

        # Salva CSVs individuais (backup)
        print("\n📁 Salvando relatórios CSV individuais...")
        salvar_csv('relatorio_email_duplicado.csv', lista_email_duplicado, headers)
        salvar_csv('relatorio_um_cpf_inexistente.csv', lista_um_inexistente, headers)
        salvar_csv('relatorio_ambos_cpf_inexistentes.csv', lista_ambos_inexistentes, headers)
        salvar_csv('relatorio_outros_erros.csv', lista_erros_outros, headers)
        
        # Salva arquivo Excel consolidado com todas as abas
        print("\n📊 Gerando arquivo Excel consolidado...")
        relatorios = {
            '1-Emails Duplicados': (lista_email_duplicado, headers),
            '2-Um CPF Inexistente': (lista_um_inexistente, headers),
            '3-Ambos CPF Inexistentes': (lista_ambos_inexistentes, headers),
            '4-Outros Erros': (lista_erros_outros, headers)
        }
        salvar_excel_consolidado(relatorios, 'relatorio_divergencias.xlsx')

if __name__ == "__main__":
    main()